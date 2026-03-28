import logging
import os
import io
from datetime import datetime, date
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, ContextTypes, filters, JobQueue
)
from database import Database, CURRENCIES, CURRENCY_SYMBOLS

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

(
    MAIN_MENU, ADD_TRANSACTION_ACCOUNT, ADD_TRANSACTION_AMOUNT,
    ADD_TRANSACTION_CATEGORY, ADD_TRANSACTION_NOTE, ADD_ACCOUNT_NAME,
    ADD_ACCOUNT_BALANCE, ADD_ACCOUNT_CURRENCY, ADD_CATEGORY_NAME,
    TRANSFER_FROM, TRANSFER_TO, TRANSFER_AMOUNT,
    SET_REMINDER_HOUR, JOIN_TEAM, CREATE_TEAM_NAME,
) = range(15)

db = Database()

# ─── KEYBOARDS ────────────────────────────────────────────────

def main_keyboard():
    return ReplyKeyboardMarkup([
        [KeyboardButton("💰 Баланс"), KeyboardButton("➕ Доход")],
        [KeyboardButton("➖ Расход"), KeyboardButton("🔄 Перевод")],
        [KeyboardButton("📊 Статистика"), KeyboardButton("📤 Экспорт")],
        [KeyboardButton("👥 Команда"), KeyboardButton("⚙️ Настройки")],
    ], resize_keyboard=True)

def accounts_keyboard(accounts, action="tr_acc"):
    sym = CURRENCY_SYMBOLS
    buttons = [[InlineKeyboardButton(
        f"{a['name']} — {float(a['balance']):,.0f} {sym.get(a.get('currency','UZS'), a.get('currency',''))}",
        callback_data=f"{action}:{a['id']}"
    )] for a in accounts]
    buttons.append([InlineKeyboardButton("❌ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(buttons)

def categories_keyboard(categories):
    buttons = []
    row = []
    for cat in categories:
        row.append(InlineKeyboardButton(f"{cat['emoji']} {cat['name']}", callback_data=f"cat:{cat['id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton("➕ Новая категория", callback_data="cat:new")])
    buttons.append([InlineKeyboardButton("❌ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(buttons)

def currency_keyboard(action="cur"):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🇺🇿 Сум (UZS)", callback_data=f"{action}:UZS"),
         InlineKeyboardButton("🇺🇸 Доллар (USD)", callback_data=f"{action}:USD")],
        [InlineKeyboardButton("🇷🇺 Рубль (RUB)", callback_data=f"{action}:RUB")],
        [InlineKeyboardButton("❌ Отмена", callback_data="cancel")],
    ])

def settings_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏦 Счета", callback_data="settings:accounts")],
        [InlineKeyboardButton("💱 Валюта по умолчанию", callback_data="settings:currency")],
        [InlineKeyboardButton("⏰ Напоминания", callback_data="settings:reminder")],
        [InlineKeyboardButton("🏷️ Категории расходов", callback_data="settings:cat_expense")],
        [InlineKeyboardButton("🏷️ Категории доходов", callback_data="settings:cat_income")],
        [InlineKeyboardButton("❌ Закрыть", callback_data="cancel")],
    ])

# ─── HELPERS ──────────────────────────────────────────────────

def fmt(amount, currency='UZS'):
    sym = CURRENCY_SYMBOLS.get(currency, currency)
    if currency == 'USD':
        return f"{float(amount):,.2f} {sym}"
    return f"{float(amount):,.0f} {sym}"

def format_balance(accounts):
    if not accounts:
        return "У вас нет счетов. Добавьте счёт через /addaccount"
    lines = ["💼 <b>Ваши счета:</b>\n"]
    by_currency = {}
    for a in accounts:
        cur = a.get('currency', 'UZS')
        lines.append(f"  <b>{a['name']}</b>: {fmt(a['balance'], cur)}")
        by_currency[cur] = by_currency.get(cur, 0) + float(a['balance'])
    lines.append("")
    for cur, total in by_currency.items():
        lines.append(f"<b>Итого {cur}: {fmt(total, cur)}</b>")
    return "\n".join(lines)

def format_stats(stats, period_name):
    if not stats:
        return f"📊 <b>Статистика за {period_name}:</b>\n\nНет транзакций."
    total_in = sum(float(t['amount']) for t in stats if t['type'] == 'income')
    total_ex = sum(float(t['amount']) for t in stats if t['type'] == 'expense')
    exp_cat = {}
    inc_cat = {}
    for t in stats:
        key = f"{t['cat_emoji'] or '📦'} {t['cat_name'] or 'Без категории'}"
        if t['type'] == 'expense':
            exp_cat[key] = exp_cat.get(key, 0) + float(t['amount'])
        else:
            inc_cat[key] = inc_cat.get(key, 0) + float(t['amount'])
    lines = [f"📊 <b>Статистика за {period_name}:</b>\n"]
    if total_in:
        lines.append(f"📈 <b>Доходы: {total_in:,.0f}</b>")
        for cat, amt in sorted(inc_cat.items(), key=lambda x: -x[1]):
            lines.append(f"   {cat}: {amt:,.0f} ({amt/total_in*100:.0f}%)")
        lines.append("")
    if total_ex:
        lines.append(f"📉 <b>Расходы: {total_ex:,.0f}</b>")
        for cat, amt in sorted(exp_cat.items(), key=lambda x: -x[1]):
            lines.append(f"   {cat}: {amt:,.0f} ({amt/total_ex*100:.0f}%)")
        lines.append("")
    net = total_in - total_ex
    lines.append(f"<b>{'➕' if net>=0 else '➖'} Итог: {net:+,.0f}</b>")
    return "\n".join(lines)

# ─── START ────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    db.ensure_user(user_id)
    await update.message.reply_text(
        f"👋 Привет, <b>{update.effective_user.first_name}</b>!\n\n"
        "Я твой финансовый менеджер 💼\n"
        "Веду учёт доходов, расходов, баланс по счетам.\n\n"
        "Начни с добавления счёта: /addaccount",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    return MAIN_MENU

# ─── BALANCE ──────────────────────────────────────────────────

async def show_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    accounts = db.get_accounts(update.effective_user.id)
    await update.message.reply_text(format_balance(accounts), parse_mode="HTML")
    return MAIN_MENU

# ─── ADD TRANSACTION ──────────────────────────────────────────

async def start_income(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tr_type'] = 'income'
    return await ask_account(update, context)

async def start_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tr_type'] = 'expense'
    return await ask_account(update, context)

async def ask_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    accounts = db.get_accounts(update.effective_user.id)
    if not accounts:
        await update.message.reply_text("Сначала добавьте счёт: /addaccount")
        return MAIN_MENU
    label = "дохода" if context.user_data['tr_type'] == 'income' else "расхода"
    await update.message.reply_text(f"Выберите счёт для {label}:", reply_markup=accounts_keyboard(accounts))
    return ADD_TRANSACTION_ACCOUNT

async def transaction_account_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    acc_id = int(query.data.split(":")[1])
    context.user_data['tr_account_id'] = acc_id
    acc = db.get_account(acc_id)
    context.user_data['tr_currency'] = acc.get('currency', 'UZS')
    await query.edit_message_text(f"Введите сумму ({CURRENCY_SYMBOLS.get(acc.get('currency','UZS'))}):")
    return ADD_TRANSACTION_AMOUNT

async def transaction_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.replace(" ", "").replace(",", "")
    try:
        amount = float(text)
        if amount <= 0: raise ValueError
    except ValueError:
        await update.message.reply_text("Введите корректную сумму, например: 50000")
        return ADD_TRANSACTION_AMOUNT
    context.user_data['tr_amount'] = amount
    cats = db.get_categories(update.effective_user.id, context.user_data['tr_type'])
    await update.message.reply_text("Выберите категорию:", reply_markup=categories_keyboard(cats))
    return ADD_TRANSACTION_CATEGORY

async def transaction_category_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    cat_id = query.data.split(":")[1]
    if cat_id == "new":
        await query.edit_message_text("Введите название новой категории:")
        return ADD_CATEGORY_NAME
    context.user_data['tr_category_id'] = int(cat_id)
    await query.edit_message_text("Добавьте комментарий (или /skip):")
    return ADD_TRANSACTION_NOTE

async def new_category_in_transaction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Название не может быть пустым."); return ADD_CATEGORY_NAME
    cat_id = db.add_category(update.effective_user.id, name, "📦", context.user_data['tr_type'])
    context.user_data['tr_category_id'] = cat_id
    await update.message.reply_text(f"Категория «{name}» создана!\n\nДобавьте комментарий (или /skip):")
    return ADD_TRANSACTION_NOTE

async def transaction_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await save_transaction(update, context, update.message.text)

async def transaction_skip_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await save_transaction(update, context, "")

async def save_transaction(update: Update, context: ContextTypes.DEFAULT_TYPE, note: str):
    user_id = update.effective_user.id
    tr_type = context.user_data['tr_type']
    account_id = context.user_data['tr_account_id']
    amount = context.user_data['tr_amount']
    category_id = context.user_data.get('tr_category_id')
    currency = context.user_data.get('tr_currency', 'UZS')
    db.add_transaction(user_id, tr_type, account_id, amount, category_id, note, currency)
    account = db.get_account(account_id)
    sign = "+" if tr_type == 'income' else "-"
    emoji = "✅📈" if tr_type == 'income' else "✅📉"
    await update.message.reply_text(
        f"{emoji} <b>Записано!</b>\n\n"
        f"{'Доход' if tr_type=='income' else 'Расход'}: <b>{sign}{fmt(amount, currency)}</b>\n"
        f"Счёт: {account['name']}\n"
        f"Остаток: <b>{fmt(account['balance'], currency)}</b>",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    context.user_data.clear()
    return MAIN_MENU

# ─── TRANSFER ─────────────────────────────────────────────────

async def start_transfer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    accounts = db.get_accounts(update.effective_user.id)
    if len(accounts) < 2:
        await update.message.reply_text("Для перевода нужно минимум 2 счёта."); return MAIN_MENU
    context.user_data['transfer'] = {}
    await update.message.reply_text("Выберите счёт <b>откуда</b>:", parse_mode="HTML",
                                     reply_markup=accounts_keyboard(accounts, "tfr_from"))
    return TRANSFER_FROM

async def transfer_from_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    acc_id = int(query.data.split(":")[1])
    context.user_data['transfer']['from'] = acc_id
    accounts = [a for a in db.get_accounts(update.effective_user.id) if a['id'] != acc_id]
    await query.edit_message_text("Выберите счёт <b>куда</b>:", parse_mode="HTML",
                                   reply_markup=accounts_keyboard(accounts, "tfr_to"))
    return TRANSFER_TO

async def transfer_to_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    context.user_data['transfer']['to'] = int(query.data.split(":")[1])
    await query.edit_message_text("Введите сумму перевода:")
    return TRANSFER_AMOUNT

async def transfer_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.replace(" ", "").replace(",", "")
    try:
        amount = float(text)
        if amount <= 0: raise ValueError
    except ValueError:
        await update.message.reply_text("Введите корректную сумму."); return TRANSFER_AMOUNT
    user_id = update.effective_user.id
    from_id = context.user_data['transfer']['from']
    to_id = context.user_data['transfer']['to']
    if not db.transfer(user_id, from_id, to_id, amount):
        await update.message.reply_text("❌ Недостаточно средств!"); return MAIN_MENU
    from_acc = db.get_account(from_id)
    to_acc = db.get_account(to_id)
    await update.message.reply_text(
        f"🔄 <b>Перевод выполнен!</b>\n\n"
        f"Сумма: <b>{amount:,.0f}</b>\n"
        f"{from_acc['name']}: {fmt(from_acc['balance'], from_acc.get('currency','UZS'))}\n"
        f"{to_acc['name']}: {fmt(to_acc['balance'], to_acc.get('currency','UZS'))}",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    context.user_data.clear()
    return MAIN_MENU

# ─── STATISTICS ───────────────────────────────────────────────

async def show_stats_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📊 Выберите период:", reply_markup=InlineKeyboardMarkup([
        [InlineKeyboardButton("Сегодня", callback_data="stats:today"),
         InlineKeyboardButton("Неделя", callback_data="stats:week")],
        [InlineKeyboardButton("Месяц", callback_data="stats:month"),
         InlineKeyboardButton("Год", callback_data="stats:year")],
        [InlineKeyboardButton("❌ Закрыть", callback_data="cancel")],
    ]))
    return MAIN_MENU

async def stats_period_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Закрыто."); return MAIN_MENU
    period = query.data.split(":")[1]
    period_names = {"today": "сегодня", "week": "неделю", "month": "месяц", "year": "год"}
    stats = db.get_stats(update.effective_user.id, period)
    await query.edit_message_text(format_stats(stats, period_names[period]), parse_mode="HTML")
    return MAIN_MENU

# ─── EXPORT ───────────────────────────────────────────────────

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        await update.message.reply_text("❌ openpyxl не установлен. Добавьте в requirements.txt")
        return MAIN_MENU

    user_id = update.effective_user.id
    transactions = db.get_all_transactions(user_id)
    accounts = db.get_accounts(user_id)

    wb = openpyxl.Workbook()

    # Sheet 1: Transactions
    ws = wb.active
    ws.title = "Транзакции"
    headers = ["Дата", "Тип", "Счёт", "Категория", "Сумма", "Валюта", "Комментарий"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 25

    for i, t in enumerate(transactions, 2):
        type_ru = "Доход" if t['type'] == 'income' else "Расход"
        cat = f"{t['cat_emoji'] or ''} {t['cat_name'] or ''}".strip()
        ws.append([
            t['created_at'].strftime("%d.%m.%Y %H:%M") if t['created_at'] else "",
            type_ru,
            t['account_name'] or "",
            cat,
            float(t['amount']),
            t.get('currency', 'UZS'),
            t['note'] or "",
        ])
        if t['type'] == 'income':
            ws.cell(row=i, column=2).font = Font(color="2E7D32")
        else:
            ws.cell(row=i, column=2).font = Font(color="C62828")

    # Sheet 2: Accounts
    ws2 = wb.create_sheet("Счета")
    ws2.append(["Счёт", "Баланс", "Валюта"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for a in accounts:
        ws2.append([a['name'], float(a['balance']), a.get('currency', 'UZS')])
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().strftime("%d-%m-%Y")
    await update.message.reply_document(
        document=buf,
        filename=f"финансы_{today}.xlsx",
        caption=f"📤 Экспорт за все время\n{len(transactions)} транзакций"
    )
    return MAIN_MENU

# ─── REMINDERS ────────────────────────────────────────────────

async def reminder_job(context: ContextTypes.DEFAULT_TYPE):
    hour = datetime.now().hour
    users = db.get_all_reminder_users(hour)
    for user in users:
        try:
            await context.bot.send_message(
                chat_id=user['id'],
                text="🔔 <b>Напоминание!</b>\n\nНе забудьте записать расходы за сегодня 💰",
                parse_mode="HTML"
            )
        except Exception as e:
            logger.warning(f"Reminder failed for {user['id']}: {e}")

async def set_reminder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "⏰ <b>Напоминания</b>\n\nВ какое время присылать напоминание каждый день?\n"
        "Введите час от 0 до 23 (по Ташкентскому времени UTC+5)\n\n"
        "Например: <code>20</code> — напоминание в 20:00\n\n"
        "Отправьте /remindoff чтобы отключить.",
        parse_mode="HTML"
    )
    return SET_REMINDER_HOUR

async def reminder_hour_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        # Convert Tashkent time (UTC+5) to UTC
        local_hour = int(text)
        if not 0 <= local_hour <= 23: raise ValueError
        utc_hour = (local_hour - 5) % 24
    except ValueError:
        await update.message.reply_text("Введите число от 0 до 23.")
        return SET_REMINDER_HOUR
    db.set_reminder(update.effective_user.id, utc_hour)
    await update.message.reply_text(
        f"✅ Напоминание установлено на <b>{text}:00</b> по Ташкенту каждый день!",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    return MAIN_MENU

async def reminder_off(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db.set_reminder(update.effective_user.id, None)
    await update.message.reply_text("🔕 Напоминания отключены.", reply_markup=main_keyboard())
    return MAIN_MENU

# ─── TEAM ─────────────────────────────────────────────────────

async def show_team(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    teams = db.get_user_teams(user_id)
    lines = ["👥 <b>Команды</b>\n"]
    if teams:
        for t in teams:
            role = "Владелец" if t['role'] == 'owner' else "Участник"
            lines.append(f"  <b>{t['name']}</b> — {t['member_count']} чел. ({role})")
            lines.append(f"  Код приглашения: <code>team_{t['id']}</code>")
            lines.append("")
    else:
        lines.append("У вас нет команд.")
    lines.append("\n/createteam — создать команду\n/jointeam — вступить по коду")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML")
    return MAIN_MENU

async def create_team_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите название команды:")
    return CREATE_TEAM_NAME

async def create_team_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Название не может быть пустым."); return CREATE_TEAM_NAME
    user_id = update.effective_user.id
    team_id = db.create_team(user_id, name)
    await update.message.reply_text(
        f"✅ Команда <b>{name}</b> создана!\n\n"
        f"Код приглашения: <code>team_{team_id}</code>\n\n"
        f"Поделитесь кодом с сотрудниками — они вступят командой /jointeam",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    return MAIN_MENU

async def join_team_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите код приглашения (например: <code>team_5</code>):", parse_mode="HTML")
    return JOIN_TEAM

async def join_team_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    code = update.message.text.strip()
    try:
        team_id = int(code.replace("team_", ""))
    except ValueError:
        await update.message.reply_text("Неверный код. Пример: team_5"); return JOIN_TEAM
    user_id = update.effective_user.id
    if db.join_team(user_id, team_id):
        await update.message.reply_text("✅ Вы вступили в команду!", reply_markup=main_keyboard())
    else:
        await update.message.reply_text("❌ Команда не найдена. Проверьте код.")
    return MAIN_MENU

# ─── ADD ACCOUNT ──────────────────────────────────────────────

async def add_account_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите название счёта (например: Наличные, Карта UzCard, Payme):")
    return ADD_ACCOUNT_NAME

async def add_account_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Название не может быть пустым."); return ADD_ACCOUNT_NAME
    context.user_data['new_account_name'] = name
    await update.message.reply_text(f"Выберите валюту для счёта «{name}»:", reply_markup=currency_keyboard("acc_cur"))
    return ADD_ACCOUNT_CURRENCY

async def add_account_currency_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    currency = query.data.split(":")[1]
    context.user_data['new_account_currency'] = currency
    sym = CURRENCY_SYMBOLS.get(currency, currency)
    await query.edit_message_text(f"Начальный баланс ({sym})? Введите 0 если пустой:")
    return ADD_ACCOUNT_BALANCE

async def add_account_balance_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.replace(" ", "").replace(",", "")
    try:
        balance = float(text)
        if balance < 0: raise ValueError
    except ValueError:
        await update.message.reply_text("Введите корректную сумму (0 или больше)."); return ADD_ACCOUNT_BALANCE
    user_id = update.effective_user.id
    name = context.user_data['new_account_name']
    currency = context.user_data.get('new_account_currency', 'UZS')
    db.add_account(user_id, name, balance, currency)
    await update.message.reply_text(
        f"✅ Счёт <b>{name}</b> создан!\nБаланс: {fmt(balance, currency)}",
        parse_mode="HTML", reply_markup=main_keyboard()
    )
    context.user_data.clear()
    return MAIN_MENU

# ─── SETTINGS ─────────────────────────────────────────────────

async def show_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚙️ <b>Настройки</b>", parse_mode="HTML", reply_markup=settings_keyboard())
    return MAIN_MENU

async def settings_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    user_id = update.effective_user.id
    if query.data == "cancel":
        await query.edit_message_text("Закрыто."); return MAIN_MENU
    action = query.data.split(":")[1]
    if action == "accounts":
        accounts = db.get_accounts(user_id)
        lines = ["🏦 <b>Ваши счета:</b>\n"] + [f"  {a['name']}: <b>{fmt(a['balance'], a.get('currency','UZS'))}</b>" for a in accounts]
        lines.append("\n/addaccount — добавить счёт")
        await query.edit_message_text("\n".join(lines), parse_mode="HTML")
    elif action == "currency":
        await query.edit_message_text("Выберите валюту по умолчанию:", reply_markup=currency_keyboard("setcur"))
    elif action == "reminder":
        user = db.get_user(user_id)
        status = f"Установлено: {(user['reminder_hour']+5)%24}:00" if user and user.get('reminder_hour') is not None else "Не установлено"
        await query.edit_message_text(
            f"⏰ <b>Напоминания</b>\nСтатус: {status}\n\n/setreminder — установить\n/remindoff — отключить",
            parse_mode="HTML"
        )
    elif action in ("cat_expense", "cat_income"):
        tp = 'expense' if action == 'cat_expense' else 'income'
        cats = db.get_categories(user_id, tp)
        lines = [f"🏷️ <b>Категории {'расходов' if tp=='expense' else 'доходов'}:</b>\n"]
        lines += [f"  {c['emoji']} {c['name']}" for c in cats]
        await query.edit_message_text("\n".join(lines), parse_mode="HTML")
    return MAIN_MENU

async def set_currency_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Отменено."); return MAIN_MENU
    currency = query.data.split(":")[1]
    db.set_currency(update.effective_user.id, currency)
    sym = CURRENCY_SYMBOLS.get(currency, currency)
    await query.edit_message_text(f"✅ Валюта по умолчанию: <b>{currency} ({sym})</b>", parse_mode="HTML")
    return MAIN_MENU

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Отменено.", reply_markup=main_keyboard())
    return MAIN_MENU

# ─── MAIN ─────────────────────────────────────────────────────

def main():
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise ValueError("Укажите BOT_TOKEN")
    app = Application.builder().token(token).build()

    # Reminder job every hour
    app.job_queue.run_repeating(reminder_job, interval=3600, first=10)

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Regex("^💰 Баланс$"), show_balance),
            MessageHandler(filters.Regex("^➕ Доход$"), start_income),
            MessageHandler(filters.Regex("^➖ Расход$"), start_expense),
            MessageHandler(filters.Regex("^🔄 Перевод$"), start_transfer),
            MessageHandler(filters.Regex("^📊 Статистика$"), show_stats_menu),
            MessageHandler(filters.Regex("^📤 Экспорт$"), export_excel),
            MessageHandler(filters.Regex("^👥 Команда$"), show_team),
            MessageHandler(filters.Regex("^⚙️ Настройки$"), show_settings),
            CommandHandler("addaccount", add_account_start),
            CommandHandler("setreminder", set_reminder),
            CommandHandler("remindoff", reminder_off),
            CommandHandler("createteam", create_team_start),
            CommandHandler("jointeam", join_team_start),
        ],
        states={
            MAIN_MENU: [
                MessageHandler(filters.Regex("^💰 Баланс$"), show_balance),
                MessageHandler(filters.Regex("^➕ Доход$"), start_income),
                MessageHandler(filters.Regex("^➖ Расход$"), start_expense),
                MessageHandler(filters.Regex("^🔄 Перевод$"), start_transfer),
                MessageHandler(filters.Regex("^📊 Статистика$"), show_stats_menu),
                MessageHandler(filters.Regex("^📤 Экспорт$"), export_excel),
                MessageHandler(filters.Regex("^👥 Команда$"), show_team),
                MessageHandler(filters.Regex("^⚙️ Настройки$"), show_settings),
                CommandHandler("addaccount", add_account_start),
                CommandHandler("setreminder", set_reminder),
                CommandHandler("remindoff", reminder_off),
                CommandHandler("createteam", create_team_start),
                CommandHandler("jointeam", join_team_start),
                CallbackQueryHandler(stats_period_selected, pattern="^stats:"),
                CallbackQueryHandler(settings_callback, pattern="^settings:"),
                CallbackQueryHandler(set_currency_callback, pattern="^setcur:"),
                CallbackQueryHandler(settings_callback, pattern="^cancel$"),
            ],
            ADD_TRANSACTION_ACCOUNT: [
                CallbackQueryHandler(transaction_account_selected, pattern="^tr_acc:|^cancel$"),
            ],
            ADD_TRANSACTION_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, transaction_amount_entered),
            ],
            ADD_TRANSACTION_CATEGORY: [
                CallbackQueryHandler(transaction_category_selected, pattern="^cat:|^cancel$"),
            ],
            ADD_CATEGORY_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, new_category_in_transaction),
            ],
            ADD_TRANSACTION_NOTE: [
                CommandHandler("skip", transaction_skip_note),
                MessageHandler(filters.TEXT & ~filters.COMMAND, transaction_note),
            ],
            ADD_ACCOUNT_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_account_name_handler),
            ],
            ADD_ACCOUNT_CURRENCY: [
                CallbackQueryHandler(add_account_currency_selected, pattern="^acc_cur:|^cancel$"),
            ],
            ADD_ACCOUNT_BALANCE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_account_balance_handler),
            ],
            TRANSFER_FROM: [
                CallbackQueryHandler(transfer_from_selected, pattern="^tfr_from:|^cancel$"),
            ],
            TRANSFER_TO: [
                CallbackQueryHandler(transfer_to_selected, pattern="^tfr_to:|^cancel$"),
            ],
            TRANSFER_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_amount_entered),
            ],
            SET_REMINDER_HOUR: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, reminder_hour_entered),
            ],
            JOIN_TEAM: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, join_team_code),
            ],
            CREATE_TEAM_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, create_team_name),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    logger.info("Бот запущен!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
